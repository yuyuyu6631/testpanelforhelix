# -*- coding: utf-8 -*-
"""
报告模块 - 生成 Excel 测试报告

实现：
1. Excel 报告生成（带样式）
2. 统计通过/失败数量
3. 按原始顺序排序
"""

import logging
import pandas as pd
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .config import Config

logger = logging.getLogger("AutoTest.Reporter")


class Reporter:
    """
    报告生成类
    
    负责：
    1. 收集测试结果
    2. 生成格式化的 Excel 报告
    3. 统计测试通过率
    """
    
    def __init__(self):
        self.results: List[Dict[str, Any]] = []
    
    def add_result(self, result: Dict[str, Any]):
        """
        添加单条测试结果
        
        Args:
            result: 测试结果字典
        """
        self.results.append(result)
    
    def generate_report(self, output_path: str = None) -> str:
        """
        生成 Excel 测试报告
        
        Args:
            output_path: 输出文件路径，默认使用配置中的路径
            
        Returns:
            str: 报告文件路径
        """
        if output_path is None:
            output_path = Config.OUTPUT_FILE
        
        if not self.results:
            logger.warning("没有测试结果，无法生成报告")
            return output_path
        
        # 按 Index 排序
        sorted_results = sorted(self.results, key=lambda x: x.get("Index", 0))
        
        # 创建 DataFrame
        df = pd.DataFrame(sorted_results)
        
        # 创建带样式的 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "测试报告"
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # 表头样式
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                # 测试结果列的样式
                elif c_idx == df.columns.get_loc("测试结果") + 1:
                    if value == "通过":
                        cell.fill = pass_fill
                    else:
                        cell.fill = fail_fill
        
        # 调整列宽
        column_widths = {
            "Index": 8,
            "问题": 40,
            "预期关键字": 20,
            "预期条件": 20,
            "实际生成的SQL": 60,
            "测试结果": 10,
            "备注": 30
        }
        
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = column_widths.get(col_name, 15)
        
        # 保存
        try:
            wb.save(output_path)
            logger.info(f"✓ 测试报告已保存: {output_path}")
        except Exception as e:
            logger.error(f"保存报告失败: {e}")
            raise
        
        return output_path
    
    def get_statistics(self) -> Dict[str, Any]:
        """
        获取测试统计信息
        
        Returns:
            Dict: 包含总数、通过数、失败数、通过率的字典
        """
        total = len(self.results)
        passed = sum(1 for r in self.results if r.get("测试结果") == "通过")
        failed = total - passed
        pass_rate = (passed / total * 100) if total > 0 else 0
        
        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "pass_rate": f"{pass_rate:.2f}%"
        }
    
    def print_summary(self):
        """打印测试摘要"""
        stats = self.get_statistics()
        
        logger.info("=" * 50)
        logger.info("测试执行完成!")
        logger.info(f"  总用例数: {stats['total']}")
        logger.info(f"  通过数量: {stats['passed']}")
        logger.info(f"  失败数量: {stats['failed']}")
        logger.info(f"  通过率: {stats['pass_rate']}")
        logger.info("=" * 50)

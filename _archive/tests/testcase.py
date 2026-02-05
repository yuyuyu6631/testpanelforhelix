import os
import time
import re
import json
import logging
import concurrent.futures
import threading
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict
from urllib.parse import urlparse

import requests
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================= Configuration =================
class Config:
    INPUT_FILE = r"D:\apiautotest\case\测试52sql预期对比.xlsx"
    OUTPUT_FILE = r"D:\apiautotest\case\测试结果报告_OOP.xlsx"
    BASE_URL = "http://52.82.11.208:8050"
    LOGIN_URL = "https://user.binarysee.com.cn/auth/login?client=global&redirecturl=http://113.44.121.105/zov-portal-web/login"
    INIT_TOKEN = "a3079926-c016-441b-b1f7-07e1f6c56816"
    MAX_WORKERS = 5  # 并发线程数
    
    HEADERS_COMMON = {
        'Accept': 'text/event-stream',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Content-Type': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
        'Tenant-Id': '1916705084137349121', 
    }

# ================= Data Models =================
@dataclass
class TestCase:
    row_idx: int
    question: str
    expected_kw: Optional[str]
    expected_cond: Optional[str]

@dataclass
class TestResult:
    case: TestCase
    actual_sql: str
    extracted_kw: str
    extracted_cond: str
    is_passed: bool
    message: str
    error: Optional[str] = None

# ================= Utils & Validator =================
class SQLValidator:
    """提供SQL解析和校验的静态方法"""
    
    @staticmethod
    def robust_sql_parse(sql: str) -> Tuple[str, str]:
        """解析SQL, 返回 (条件内容, 字段内容)"""
        if pd.isna(sql) or not isinstance(sql, str):
            return "", ""
        
        # 简化版正则解析
        sql = sql.strip().replace('\n', ' ')
        
        # 1. 提取 WHERE 子句
        where_match = re.search(r'WHERE\s+(.+?)(?:\s+(?:GROUP BY|ORDER BY|LIMIT|HAVING)|$)', sql, re.IGNORECASE)
        where_str = where_match.group(1) if where_match else ""
        
        # 2. 提取 SELECT 字段
        select_match = re.search(r'SELECT\s+(.+?)\s+FROM', sql, re.IGNORECASE)
        fields_str = select_match.group(1) if select_match else ""
        
        extracted_conds = []
        if where_str:
            # 分割 AND/OR
            conds = re.split(r'\s+(?:AND|OR)\s+', where_str, flags=re.IGNORECASE)
            for cond in conds:
                cond = cond.strip()
                # 提取 字段 op 值
                op_pattern = r'\s*(=|!=|<>|>=|<=|>|<|\bLIKE\b|\bIN\b|\bBETWEEN\b)\s*'
                parts = re.split(op_pattern, cond, maxsplit=1, flags=re.IGNORECASE)
                if len(parts) >= 3:
                    field = parts[0].strip()
                    val = parts[2].strip()
                    # 去除引号
                    val = re.sub(r'^[\'"`](.*)[\'"`]$', r'\1', val)
                    extracted_conds.append(f"{field}\n{val}")

        extracted_fields = []
        if fields_str:
            # 简单按逗号分割，暂不处理复杂嵌套括号
            fields = fields_str.split(',')
            for f in fields:
                f = f.strip()
                # 尝试提取 AS 别名
                alias_match = re.search(r'(.+?)\s+(?:AS\s+)?([`\'"]?[\w\u4e00-\u9fa5]+[`\'"]?)$', f, re.IGNORECASE)
                if alias_match:
                    expr = alias_match.group(1).strip()
                    alias = alias_match.group(2).strip().strip('`\'"')
                    if '(' not in alias:
                        extracted_fields.append(f"{expr}\n{alias}")

        return "\n\n".join(extracted_conds), "\n\n".join(extracted_fields)

    @staticmethod
    def validate(extracted_kw: str, extracted_cond: str, expected_kw: str, expected_cond: str) -> Tuple[bool, str]:
        """校验逻辑"""
        if not expected_kw and not expected_cond:
            return True, "无预期答案"
            
        def parse_expectations(text):
            if not text: return []
            # 支持 'or' 分组
            groups = text.split('\nor\n') if '\nor\n' in str(text) else [str(text)]
            return [[item.strip().lower() for item in g.split('\n') if item.strip()] for g in groups]

        def get_actual_set(text):
            if not text: return set()
            lines = str(text).split('\n')
            res = set()
            for line in lines:
                line = line.strip()
                # 排除字段前缀 a., h. 等
                if line and not re.match(r'^[a-z_]+\.', line, re.I):
                    res.add(line.lower())
            return res

        # 1. 校验关键字
        kw_groups = parse_expectations(expected_kw)
        actual_kw_set = get_actual_set(extracted_kw)
        
        kw_pass = False
        missing_kw_Log = []
        
        if not expected_kw:
            kw_pass = True
        else:
            for idx, group in enumerate(kw_groups):
                missing = [k for k in group if k not in actual_kw_set]
                if not missing:
                    kw_pass = True
                    break
                if idx == 0: missing_kw_Log = missing # 记录第一组缺失的作为参考
        
        # 2. 校验条件
        cond_groups = parse_expectations(expected_cond)
        actual_cond_set = get_actual_set(extracted_cond)
        
        cond_pass = False
        missing_cond_log = []
        
        if not expected_cond:
            cond_pass = True
        else:
            for idx, group in enumerate(cond_groups):
                missing = [c for c in group if c not in actual_cond_set]
                if not missing:
                    cond_pass = True
                    break
                if idx == 0: missing_cond_log = missing

        if kw_pass and cond_pass:
            return True, "通过"
        
        reasons = []
        if not kw_pass: reasons.append(f"KW缺失:{','.join(missing_kw_Log)}")
        if not cond_pass: reasons.append(f"COND缺失:{','.join(missing_cond_log)}")
        
        return False, "未通过: " + "; ".join(reasons)

# ================= Auth & Client =================
class AuthManager:
    """负责Token的获取和刷新"""
    def __init__(self):
        self._token = Config.INIT_TOKEN
        self._lock = threading.Lock()

    def get_token(self):
        with self._lock:
            return self._token

    def update_token(self, new_token):
        with self._lock:
            self._token = new_token
            logger.info(f"Token updated: {new_token[:10]}...")

    def login_and_refresh(self) -> bool:
        """使用Playwright进行登录"""
        logger.info("Starting Auto-Login process...")
        try:
            from playwright.sync_api import sync_playwright
            
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                
                logger.info(f"Navigating to {Config.LOGIN_URL}")
                page.goto(Config.LOGIN_URL)
                
                # 等待输入框
                page.wait_for_selector("input[type='password']", timeout=15000)
                
                # 输入账号密码
                page.locator("input[placeholder*='账号']").or_(page.locator("input[type='text']").first).fill("13439427048")
                page.locator("input[type='password']").fill("Zsh@417418")
                
                # 点击登录
                page.get_by_text("登录", exact=True).click()
                
                # 等待跳转
                logger.info("Waiting for redirect...")
                try:
                    page.wait_for_url(lambda u: "zov-portal-web" in u, timeout=20000)
                except:
                    logger.warning("Redirection timeout, checking storage anyway.")
                
                # 获取 Storage
                storage_str = page.evaluate("() => localStorage.getItem('zov-user-storage')")
                browser.close()
                
                if storage_str:
                    data = json.loads(storage_str)
                    token = data.get('token')
                    if token:
                        self.update_token(token)
                        return True
            return False
            
        except Exception as e:
            logger.error(f"Login failed: {e}")
            return False

class APIClient:
    """封装API请求，自动处理重试和认证"""
    
    def __init__(self, auth_manager: AuthManager):
        self.auth = auth_manager
        self.session = requests.Session()
        self.session.headers.update(Config.HEADERS_COMMON)

    def ask_ai(self, question: str) -> str:
        """提问并返回 SQL (String)"""
        url = f"{Config.BASE_URL}/brain/faq/session/ask"
        
        payload = {
            "robot_id": "89040136787200", 
            "robotBizId": 89040136787201,
            "incremental": False,
            "status": "1",
            "chatId": 76905036366157, # 这里最好动态生成或获取，暂时沿用常量
            "user_input": question,
            "newChatId": None,
            "pageNum": 1,
            "pageSize": 10,
            "appId": "0"
        }
        
        max_retries = 2
        for attempt in range(max_retries + 1):
            current_token = self.auth.get_token()
            headers = {"Authorization": current_token}
            
            try:
                resp = self.session.post(url, json=payload, headers=headers, stream=True, timeout=30)
                
                if resp.status_code == 401:
                    logger.warning("Received 401 Unauthorized.")
                    if attempt < max_retries:
                        if self.auth.login_and_refresh():
                            continue # Retry with new token
                        else:
                            return "Error: Login Failed"
                    else:
                        return "Error: 401 Unauthorized (Max retries)"
                
                if resp.status_code != 200:
                    return f"Error: HTTP {resp.status_code}"
                
                # 处理流式响应，提取 SQL
                full_content = ""
                for line in resp.iter_lines():
                    if line:
                        full_content += line.decode('utf-8')
                        
                # 正则提取 SQL
                # 模式1: 【最终SQL校验】：SELECT ...
                match = re.search(r"【最终SQL校验】：(SELECT .+?)(?:\\n|\"|$|\\r)", full_content, re.IGNORECASE)
                if not match:
                    # 模式2: "sql":"SELECT ..."
                    match = re.search(r'"sql"\s*:\s*"(SELECT .+?)"', full_content, re.IGNORECASE)
                
                if match:
                    sql_str = match.group(1)
                    # 清理转义
                    return sql_str.replace('\\"', '"').replace('\\n', ' ').strip()
                
                return "Error: No SQL found in response"
                
            except Exception as e:
                if attempt < max_retries:
                    time.sleep(1)
                    continue
                return f"Error: Request Exception {str(e)}"
                
        return "Error: Unknown"

# ================= Data & Runner =================
class ExcelManager:
    """处理Excel读写"""
    
    @staticmethod
    def load_test_cases(file_path: str) -> List[TestCase]:
        cases = []
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            return cases
            
        wb = openpyxl.load_workbook(file_path)
        # 假设只处理第一个 Sheet
        sheet = wb.active 
        
        # 定位列
        header = [c.value for c in list(sheet.iter_rows(min_row=1, max_row=1))[0]]
        col_map = {}
        for idx, val in enumerate(header):
            if not val: continue
            v = str(val).strip()
            if "问题" in v or "question" in v.lower(): col_map['q'] = idx
            elif "预期关键字" in v: col_map['ek'] = idx
            elif "预期条件" in v: col_map['ec'] = idx
            
        if 'q' not in col_map:
            logger.error("Could not find '问题' column.")
            return cases
            
        for r_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            q_val = row[col_map['q']]
            if not q_val: continue
            
            ek_val = row[col_map['ek']] if 'ek' in col_map else None
            ec_val = row[col_map['ec']] if 'ec' in col_map else None
            
            cases.append(TestCase(
                row_idx=r_idx,
                question=str(q_val).strip(),
                expected_kw=str(ek_val).strip() if ek_val else None,
                expected_cond=str(ec_val).strip() if ec_val else None
            ))
            
        return cases

    @staticmethod
    def save_report(results: List[TestResult], output_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Report"
        
        # Writing Header
        headers = ["序号", "问题", "提取的SQL", "预期的关键字", "实际的关键字", "预期的条件", "实际的条件", "测试通过情况"]
        ws.append(headers)
        
        # Style
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            
        for res in results:
            row_data = [
                res.case.row_idx,
                res.case.question,
                res.actual_sql,
                res.case.expected_kw,
                res.extracted_kw,
                res.case.expected_cond,
                res.extracted_cond,
                res.message
            ]
            ws.append(row_data)
        
        try:
            wb.save(output_path)
            logger.info(f"Report saved to {output_path}")
        except Exception as e:
            logger.error(f"Failed to save report: {e}")

class TestEngine:
    def __init__(self):
        self.auth_manager = AuthManager()
        self.client = APIClient(self.auth_manager)
        self.results = []
        
    def process_case(self, case: TestCase) -> TestResult:
        logger.info(f"Running Row {case.row_idx}: {case.question[:20]}...")
        
        sql = self.client.ask_ai(case.question)
        
        ext_cond, ext_kw = "", ""
        is_passed = False
        msg = ""
        
        if not sql.startswith("Error"):
            ext_cond, ext_kw = SQLValidator.robust_sql_parse(sql)
            is_passed, msg = SQLValidator.validate(ext_kw, ext_cond, case.expected_kw, case.expected_cond)
        else:
            msg = sql # Error message
            
        return TestResult(
            case=case,
            actual_sql=sql,
            extracted_kw=ext_kw,
            extracted_cond=ext_cond,
            is_passed=is_passed,
            message=msg
        )

    def run(self, limit: Optional[int] = None):
        logger.info("loading test cases...")
        cases = ExcelManager.load_test_cases(Config.INPUT_FILE)
        if not cases:
            logger.error("No cases found.")
            return

        if limit:
            logger.info(f"Limiting to first {limit} cases as requested.")
            cases = cases[:limit]

        logger.info(f"Found {len(cases)} cases. Starting execution with {Config.MAX_WORKERS} workers.")
        
        completed_count = 0
        total_count = len(cases)
        results = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
            future_to_case = {executor.submit(self.process_case, case): case for case in cases}
            
            for future in concurrent.futures.as_completed(future_to_case):
                try:
                    res = future.result()
                    results.append(res)
                    completed_count += 1
                    if completed_count % 5 == 0:
                        logger.info(f"Progress: {completed_count}/{total_count} ({completed_count/total_count:.1%})")
                except Exception as e:
                    logger.error(f"Case failed: {e}")
                    
        # Sort by row index
        results.sort(key=lambda x: x.case.row_idx)
        
        # Stats
        passed = sum(1 for r in results if r.is_passed)
        rate = passed / total_count if total_count > 0 else 0
        logger.info("="*50)
        logger.info(f"Test Completed.")
        logger.info(f"Total: {total_count}, Passed: {passed}, Rate: {rate:.2%}")
        logger.info("="*50)
        
        # Save
        ExcelManager.save_report(results, Config.OUTPUT_FILE)

if __name__ == "__main__":
    import sys
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else None
    engine = TestEngine()
    engine.run(limit)
